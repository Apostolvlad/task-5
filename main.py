
import re

from openpyxl import Workbook, load_workbook

TABLE_NAME = 'remont1.xlsx' # откуда брать фразы


def check_ids(row):
    if row[1] is None or row[3] is None: return
    word = re.sub(' \w{1,3} ', ' ', row[1].lower().replace(' в минске', '').replace(' минск', '').replace('-', ' '))
    return (word, int(row[3]))

def load_ids():
    wb = load_workbook('table.xlsx')
    ws = wb['ID container']
    r = ws.values
    next(r)
    return dict(filter(lambda x: x, map(check_ids, r)))

def process_phrase():
    wb = load_workbook(TABLE_NAME)
    base_result = list()
    for sheet in wb._sheets:
        r = sheet.values
        for row in r:
            if row[0] != 'Phrase': continue
            result = list()
            word = re.sub(' \w{1,3} ', ' ', next(r)[0].lower().replace(' в минске', '').replace(' минск', '').replace('-', ' '))
            result.append(word)
            for row in r:
                if row[1] == 'Шаблон мета': break
            result.append(next(r)[2])
            result.append(next(r)[2])
            base_result.append(result)
    return base_result

def main():
    base_ids = load_ids()
    base_phrase = process_phrase()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Результат'
    ws.append(['Phrase', 'Title RS', 'Description RS', 'ID container', 'template'])
    for info in base_phrase:
        if not base_ids.get(info[0]):
            for ids in base_ids.items():
                if ids[0].startswith(info[0]):
                    info.append(ids[1])
                    break
            else:
                print(info[0])
        else:
            info.append(base_ids.get(info[0]))

        info.append(f"update SeoData set title = '{info[1]} ', description = '{info[2]}', seoStory = '' where id = (select seoId from Vacancy where id = {info[3]});")
        ws.append(info)
    wb.save("result.xlsx")

if __name__ == '__main__':
    main()
    